# This Python file uses the following encoding: utf-8

from openpyxl import load_workbook

class Excel():

    wb = load_workbook(filename = 'documents/excel-input/evaluations.xlsx', data_only=True)

    def get_semester_grades(self, sheet_name):

        # initialize variable for the grades
        grades = {}

        # get sheet from excel file with the name passed
        ws = self.wb[sheet_name]

        # get total number of rows of that sheet
        total_rows = ws.get_highest_row() - 1

        # iterate over rows
        for row_i in range(2, total_rows):

            # get student name
            student_name = None
            try:
                student_name = ws['C%s' % row_i].value.encode('utf8').strip()
                # initialize student on the grades object
                grades[student_name] = {}
                grades[student_name]['total_grade'] = float(ws['O%s' % row_i].value)
                grades[student_name]['teacher_name'] = ws['A2'].value.encode('utf8').strip()
                grades[student_name]['area_name'] = ws['D2'].value.encode('utf8').strip()
                grades[student_name]['trimester_number'] = sheet_name
                grades[student_name]['grades'] = {}
                
                # iterate over chriteria
                for col_i in xrange(ord('H'), ord('M')+1):

                    # get chriteria name
                    chriteria = ws['%s1' % chr(col_i)].value

                    # get chriteria grade value
                    grade_value = ws['%s%s' % (chr(col_i), row_i)].value

                    grades[student_name]['grades'][chriteria] = grade_value
            except:
                continue

        return grades
